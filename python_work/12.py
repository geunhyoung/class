from openpyxl import Workbook,load_workbook
import random

#workbook 파일 새로 생성
#load_workbook 파일 열기
#random 사용하면 랜덤 값 가져오기

'''
    A10~ J10
    1행 1열 10행10열까지 랜덤한 값을 넣어서
    random.xlsx 로 저장하고
    random.xlsx 불러와서
    1행열 ~10행10열까지 출력하기

    문제
    star.xlxs 파일 저장하기 
    *
    * *
    * * *
    * * * *
    * * * * *
    star.xlxs 파일 불러오기
    *
    * *
    * * *
    * * * *
    * * * * *
    클래스명 star 로 작성하기
    1. 메서드 makeStar()
    2. 메서드 loadStar()

    

'''
wb= Workbook()
ws = wb.active
#A10~ J10
#1행 1열 10행10열까지 랜덤한 값을 넣어서
for x in range(1,6):
    for y in range(1,6):
        if(y == 1):
            if(x<2):
                ws.cell(row=y,column=x).value = "*"
        if(y==2):
            if(x<3):
                ws.cell(row=y,column=x).value = "*"
        if(y==3):
            if(x<4):
                ws.cell(row=y,column=x).value = "*"
        if(y==4):
            if(x<5):
                ws.cell(row=y,column=x).value = "*"
        if(y==5):
            if(x<6):
                ws.cell(row=y,column=x).value = "*"
                
        

wb.save("star.xlsx")
wb.close()

#불러와서 출력하기
lb= load_workbook("star.xlsx")
for x in range(1,6):
    for y in range(1,6):  
        print(ws.cell(row=y,column=x).value,end=" ")
    print()
lb.close()

