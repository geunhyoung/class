from openpyxl import Workbook,load_workbook
import random

#workbook 파일 새로 생성
#load_workbook 파일 열기
#random 사용하면 랜덤 값 가져오기

'''
    A10~ J10
    1행 1열 10행10열까지 랜덤한 값을 넣어서
    random.xlsx 로 저장하고
    random.xlsx 불러와서
    1행열 ~10행10열까지 출력하기

    문제
    star.xlxs 파일 저장하기 
    *
    * *
    * * *
    * * * *
    * * * * *
    star.xlxs 파일 불러오기
    *
    * *
    * * *
    * * * *
    * * * * *
    클래스명 star 로 작성하기
    1. 메서드 makeStar()
    2. aptjem loadStar()

    

'''
wb= Workbook()
ws = wb.active
#A10~ J10
#1행 1열 10행10열까지 랜덤한 값을 넣어서
for x in range(1,11):
    for y in range(1,11):
        ws.cell(row=y,column=x).value = random.randint(0,100)

wb.save("random.xlsx")
wb.close()

#불러와서 출력하기
lb= load_workbook("random.xlsx")
for x in range(1,11):
    for y in range(1,11):  
        print(ws.cell(row=y,column=x).value,end=" ")
    print()
lb.close()

