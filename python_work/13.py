from openpyxl import Workbook,load_workbook
import random

class MyWorkbook:

    def doSave(self):
        wb= Workbook()
        ws = wb.active
        #A10~ J10
        #1행 1열 10행10열까지 랜덤한 값을 넣어서
        for x in range(1,11):
            for y in range(1,11):
                ws.cell(row=y,column=x).value = random.randint(0,100)

        wb.save("star.xlsx")
        wb.close()
#불러와서 출력하기
    def doLoad(self):
        lb= load_workbook("star.xlsx")
        ws= lb.active
        for x in range(1,11):
            for y in range(1,11):  
                print(ws.cell(row=y,column=x).value,end=" ")
            print()
        lb.close()