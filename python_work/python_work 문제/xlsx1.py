from openpyxl import Workbook,load_workbook
import random

class MyWorkbook:

    def makeStar(self):
        wb= Workbook()
        ws = wb.active
        #A10~ J10
        #1행 1열 10행10열까지 랜덤한 값을 넣어서
        for x in range(1,6):
            for y in range(1,x+1):
                
                ws.cell(row=x,column=y).value = "*"

        wb.save("star.xlsx")
        wb.close()
#불러와서 출력하기
    def loadStar(self):
        lb= load_workbook("star.xlsx")
        ws= lb.active
        for x in range(1,6):
            for y in range(1,x+1):  
                print(ws.cell(row=x,column=y).value,end=" ")
            print()
        lb.close()