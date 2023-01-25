from selenium import webdriver
from selenium.webdriver.common.by import By
import time, random
import openpyxl


def chrome_driver():
    options = webdriver.ChromeOptions()
    # options.add_argument("headless")  # 웹 브라우저를 시각적으로 띄우지 않는 headless chrome 옵션
    options.add_experimental_option('detach', True)
    options.add_experimental_option("excludeSwitches", ["enable-logging"])  # 개발도구 로그 숨기기
    user_agent = 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/100.0.4896.75 Safari/537.36'
    options.add_argument(f'user-agent={user_agent}')
    options.add_argument("start-maximized") # 창 크기 최대로
    options.add_argument("disable-infobars")
    options.add_argument("--disable-extensions")
    driver = webdriver.Chrome('./chromedriver.exe', options=options)
    return driver


def shopping_list(keyword):

    ## 추출한 데이터를 저장할 엑셀 파일(검색 상품명) 생성
    fileName = keyword + '.xlsx'
    wb = openpyxl.Workbook()
    # wb.save('./' + fileName)
    ws = wb['Sheet']
    theme = ['검색 상품명', '상품명', '최저', '가격', '카테고리', '리뷰', '구매건수', '등록월', '업체명', '등급', '최저가']
    ws.append(theme)
    wb.save('./' + fileName)

    driver = chrome_driver()

    url = f'https://search.shopping.naver.com/search/all?query={keyword}&frm=NVSHATC'
    driver.get(url)
    time.sleep(1.5)
    
    ## 네이버 쇼핑은 마우스 스크롤 할 때마다 리스트가 추가되기때문에 페이지 맨 마지막까지 스크롤 시킴
    for i in range(7):
        driver.execute_script('window.scrollTo(0, document.body.scrollHeight);')
        time.sleep(random.uniform(0.3, 0.7))

    # 상품 리스트 : ul 태그 밑에 li 태그로 되어있음
    ul_tag = driver.find_element(By.CLASS_NAME, 'list_basis')
    lis = ul_tag.find_elements(By.CLASS_NAME, 'basicList_item__2XT81')
    print(len(lis))
    
    rcnt = 2
    for li in lis:
        
        div_title = li.find_element(By.CLASS_NAME, 'basicList_title__3P9Q7')
        title = div_title.text.strip()

        if len(li.find_elements(By.CLASS_NAME, 'ad_ad_stk__12U34')):
            title = 'AD_' + title  # 광고노출상품
        print(title)

        lower = ''
        if len(li.find_elements(By.CLASS_NAME, 'price_low__2vp2A')):
            lower = li.find_element(By.CLASS_NAME, 'price_low__2vp2A').text
        price = li.find_element(By.CLASS_NAME, 'price_num__2WUXn').text.replace('원', '').replace(',', '')
        if lower:
            print(lower, price)
        else:
            print(price)

        category_div = li.find_element(By.CLASS_NAME, 'basicList_depth__2QIie')
        category_span = category_div.find_elements(By.TAG_NAME, 'span')
        
        category = ''
        for cat_span in category_span:
            category += cat_span.text + '>'
        print(category)
        
        ## 리뷰,별점,구매건수,등록일 관련 parent element
        rev_reg = li.find_element(By.CLASS_NAME, 'basicList_etc_box__1Jzg6')
        
        # a 태그의 텍스트로 분기해서 리뷰,구매건수 추출
        a_tag = rev_reg.find_elements(By.TAG_NAME, 'a')
        #print(len(a_tag))

        review = ''
        purch_cnt = ''
        for a in a_tag:
            if a.text[:2] == '리뷰':
                review = a.text
            if a.text[:4] == '구매건수':
                purch_cnt = a.text[4:]

        # span 태그의 텍스트로 등록일 추출
        span_tag = rev_reg.find_elements(By.TAG_NAME, 'span')
        regist_day = ''
        for span in span_tag:
            if span.text[:3] == '등록일':
                regist_day = span.text[4:]

        print(f'리뷰:{review}, 구매건수:{purch_cnt}, 등록일:{regist_day}')
    
        ## 쇼핑몰 등급
        lower_list = ''
        if len(li.find_elements(By.CLASS_NAME, 'price_low__2vp2A')):  # '최저' 문구 있으면 - 상품카타로그 묶임 부분
            lis = li.find_elements(By.CSS_SELECTOR, 'div.basicList_mall_area__lIA7R > ul > li')  # 업체와 가격 리스트
            # print(len(lis))
            for li in lis:
                lo_cmpy = li.find_element(By.CLASS_NAME, 'basicList_mall_name__1XaKA').text
                lo_price = li.find_element(By.CLASS_NAME, 'basicList_price__2r23_').text
                lower_list += lo_cmpy + ':' + lo_price + ', '
            print(f'{lower_list}\n\n')

        else:  # 상품카타로그 묶임 아닌 경우
            company = ''
            if cmpy := li.find_element(By.CLASS_NAME, 'basicList_mall__sbVax').text:  # a 태그에 업체명이 있는 경우
                company = cmpy

            else:  # a 태그에 업체명이 없고 이미지인 경우(오픈마켓 명칭 등), img alt로 찾음
                company = li.find_element(By.CLASS_NAME, 'basicList_mall__sbVax').find_element(By.TAG_NAME, 'img').get_attribute('alt')
            
            # 업체 등급
            div_grade = ''
            try:
                div_grade = li.find_element(By.CLASS_NAME, 'basicList_grade__LMHXE').text
            except:
                print('무등급!')
            print(f'{company}: {div_grade}\n\n')

        ## 엑셀에 저장
        ws.cell(rcnt, 1).value = keyword
        ws.cell(rcnt, 2).value = title
        ws.cell(rcnt, 3).value = lower
        ws.cell(rcnt, 4).value = price
        ws.cell(rcnt, 5).value = category
        ws.cell(rcnt, 6).value = review
        ws.cell(rcnt, 7).value = purch_cnt
        ws.cell(rcnt, 8).value = regist_day
        ws.cell(rcnt, 9).value = company
        ws.cell(rcnt, 10).value = div_grade
        ws.cell(rcnt, 11).value = lower_list

        wb.save('./' + fileName)               

        rcnt += 1 
    
    driver.close()

shopping_list('샷시')
